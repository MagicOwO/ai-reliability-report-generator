import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
import logging

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """Generates Excel reports from incident data with AI-powered analysis"""

    def __init__(self, output_path: str):
        self.output_path = output_path

    def generate_report(self, 
                       incidents: List[Dict[str, Any]], 
                       analysis_results: Dict[str, Any],
                       ai_analysis: Dict[str, Any] = None):
        """
        Generate Excel report with multiple sheets
        
        Args:
            incidents: List of incident dictionaries
            analysis_results: Dictionary containing analysis results
            ai_analysis: Dictionary containing AI-powered analysis (optional)
        """
        logger.info("Starting Excel report generation")
        
        # Create Excel writer
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl')
        
        # Generate basic incident sheet
        self._generate_incident_sheet(incidents, writer)
        
        # Generate category sheet
        self._generate_category_sheet(analysis_results, writer)
        
        # Add AI analysis sheets if available
        if ai_analysis:
            logger.info("Adding AI analysis to Excel report")
            self._generate_ai_category_sheet(ai_analysis, writer)
            self._generate_key_issues_sheet(ai_analysis, writer)
            self._generate_comparative_sheet(ai_analysis, writer)
        
        # Save the Excel file
        writer.close()
        
        # Add charts and formatting with openpyxl
        self._add_charts_and_formatting()
        
        logger.info(f"Excel report saved to {self.output_path}")

    def _generate_incident_sheet(self, incidents: List[Dict[str, Any]], writer: pd.ExcelWriter):
        """Generate sheet with all incidents"""
        logger.info("Generating incidents sheet")
        
        # Extract data to pandas DataFrame
        incidents_data = []
        for incident in incidents:
            # Convert datetime to string if needed
            incident_date = incident.get('date')
            if isinstance(incident_date, datetime):
                date_str = incident_date.strftime('%Y-%m-%d %H:%M:%S')
            else:
                date_str = str(incident_date)
                
            # Get AI category if available, otherwise use standard category
            category = incident.get('ai_category', incident.get('category', 'Uncategorized'))
            
            # Extract root cause if available
            root_cause = incident.get('root_cause', 'Unknown')
            
            # Get severity if available
            severity = incident.get('severity', 'Unknown')
            
            # Create row
            row = {
                'Company': incident.get('company', ''),
                'Date': date_str,
                'Title': incident.get('title', ''),
                'Description': incident.get('description', ''),
                'Status': incident.get('status', ''),
                'Duration': incident.get('duration', 'Unknown'),
                'Category': category,
                'Severity': severity,
                'Root Cause': root_cause
            }
            incidents_data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(incidents_data)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Incidents', index=False)
        
        # Get the sheet to adjust column widths
        worksheet = writer.sheets['Incidents']
        
        # Adjust column widths
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).apply(len).max(), len(col) + 2)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 50)

    def _generate_category_sheet(self, analysis_results: Dict[str, Any], writer: pd.ExcelWriter):
        """Generate sheet with category analysis"""
        logger.info("Generating categories sheet")
        
        categories = analysis_results.get('categories', {})
        category_data = []
        
        for category, data in categories.items():
            row = {
                'Category': category,
                'Count': data.get('count', 0),
                'Percentage': data.get('percentage', 0),
            }
            category_data.append(row)
            
        # Create DataFrame
        df = pd.DataFrame(category_data)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Categories', index=False)
        
        # Get the sheet to adjust column widths
        worksheet = writer.sheets['Categories']
        
        # Adjust column widths
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).apply(len).max(), len(col) + 2)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 50)

    def _generate_ai_category_sheet(self, ai_analysis: Dict[str, Any], writer: pd.ExcelWriter):
        """Generate sheet with AI category analysis"""
        logger.info("Generating AI categories sheet")
        
        # Get categories from AI analysis
        categories = ai_analysis.get('target_analysis', {}).get('categories', [])
        
        if not categories:
            logger.warning("No AI categories found, skipping AI category sheet")
            return
        
        category_data = []
        for cat in categories:
            row = {
                'Category': cat.get('name', 'Unknown'),
                'Description': cat.get('description', ''),
                'Example': cat.get('example_incident', '')
            }
            category_data.append(row)
            
        # Create DataFrame
        df = pd.DataFrame(category_data)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='AI Categories', index=False)
        
        # Get the sheet to adjust column widths
        worksheet = writer.sheets['AI Categories']
        
        # Adjust column widths - descriptions can be long
        worksheet.column_dimensions['A'].width = 25  # Category name
        worksheet.column_dimensions['B'].width = 60  # Description
        worksheet.column_dimensions['C'].width = 50  # Example

    def _generate_key_issues_sheet(self, ai_analysis: Dict[str, Any], writer: pd.ExcelWriter):
        """Generate sheet with key issues identified by AI"""
        logger.info("Generating key issues sheet")
        
        # Get key issues from AI analysis
        key_issues = ai_analysis.get('target_analysis', {}).get('key_issues', [])
        
        if not key_issues:
            logger.warning("No key issues found, skipping key issues sheet")
            return
        
        issues_data = []
        for issue in key_issues:
            row = {
                'Issue': issue.get('title', 'Unknown'),
                'Impact': issue.get('impact', 'Unknown'),
                'Description': issue.get('description', ''),
                'Frequency': issue.get('frequency', 'Unknown')
            }
            issues_data.append(row)
            
        # Create DataFrame
        df = pd.DataFrame(issues_data)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Key Issues', index=False)
        
        # Get the sheet to adjust column widths
        worksheet = writer.sheets['Key Issues']
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 30  # Issue title
        worksheet.column_dimensions['B'].width = 15  # Impact
        worksheet.column_dimensions['C'].width = 60  # Description
        worksheet.column_dimensions['D'].width = 15  # Frequency

    def _generate_comparative_sheet(self, ai_analysis: Dict[str, Any], writer: pd.ExcelWriter):
        """Generate sheet with comparative analysis"""
        logger.info("Generating comparative analysis sheet")
        
        # Get comparative analysis
        comparative = ai_analysis.get('comparative_analysis', {})
        
        if not comparative:
            logger.warning("No comparative analysis found, skipping comparative sheet")
            return
        
        # Create summary data
        summary_data = []
        
        if 'summary' in comparative:
            summary_data.append({'Item': 'Summary', 'Value': comparative['summary']})
        
        if 'trend_comparison' in comparative:
            summary_data.append({'Item': 'Trend Comparison', 'Value': comparative['trend_comparison']})
        
        # Add common categories
        common_cats = comparative.get('common_categories', [])
        if common_cats:
            summary_data.append({'Item': 'Common Categories', 'Value': ', '.join(common_cats)})
        
        # Add unique categories
        target_unique = comparative.get('target_unique_categories', [])
        if target_unique:
            summary_data.append({'Item': 'Target Unique Categories', 'Value': ', '.join(target_unique)})
        
        peer_unique = comparative.get('peer_unique_categories', [])
        if peer_unique:
            summary_data.append({'Item': 'Peer Unique Categories', 'Value': ', '.join(peer_unique)})
        
        # Create DataFrame
        df = pd.DataFrame(summary_data)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Comparative Analysis', index=False)
        
        # Get the sheet to adjust column widths
        worksheet = writer.sheets['Comparative Analysis']
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 25  # Item
        worksheet.column_dimensions['B'].width = 80  # Value

    def _add_charts_and_formatting(self):
        """Add charts and formatting to the Excel file"""
        logger.info("Adding charts and formatting to Excel report")
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(self.output_path)
            
            # Format the incidents sheet
            if 'Incidents' in wb.sheetnames:
                sheet = wb['Incidents']
                self._format_table_header(sheet)
                
                # Add conditional formatting for severity
                self._add_severity_formatting(sheet)
            
            # Add category charts
            if 'Categories' in wb.sheetnames:
                sheet = wb['Categories']
                self._format_table_header(sheet)
                
                # Add a pie chart for categories
                self._add_category_pie_chart(sheet)
            
            # Format AI category sheet
            if 'AI Categories' in wb.sheetnames:
                sheet = wb['AI Categories']
                self._format_table_header(sheet)
                
                # Make description column text wrap
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
                    sheet.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
                    
                # Adjust row height
                for row in range(2, sheet.max_row + 1):
                    sheet.row_dimensions[row].height = 100
            
            # Format key issues sheet
            if 'Key Issues' in wb.sheetnames:
                sheet = wb['Key Issues']
                self._format_table_header(sheet)
                
                # Make description column text wrap
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
                    
                # Adjust row height
                for row in range(2, sheet.max_row + 1):
                    sheet.row_dimensions[row].height = 80
                    
                # Add impact color formatting
                self._add_impact_formatting(sheet)
            
            # Format comparative analysis sheet
            if 'Comparative Analysis' in wb.sheetnames:
                sheet = wb['Comparative Analysis']
                self._format_table_header(sheet)
                
                # Make value column text wrap
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
                    
                # Adjust row height
                for row in range(2, sheet.max_row + 1):
                    sheet.row_dimensions[row].height = 80
            
            # Save the workbook
            wb.save(self.output_path)
            logger.info("Charts and formatting added to Excel report")
            
        except Exception as e:
            logger.error(f"Error adding charts and formatting: {str(e)}")

    def _format_table_header(self, sheet):
        """Format the header row of a table"""
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply styles to header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _add_severity_formatting(self, sheet):
        """Add conditional formatting for severity column"""
        # Find severity column index
        severity_col = None
        for i, cell in enumerate(sheet[1]):
            if cell.value == 'Severity':
                severity_col = i + 1
                break
        
        if severity_col is None:
            return
        
        # Define styles
        critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        major_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")    # Orange
        minor_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")    # Yellow
        low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")      # Green
        
        # Apply styles
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=severity_col)
            value = cell.value.lower() if cell.value else ""
            
            if 'critical' in value:
                cell.fill = critical_fill
            elif 'major' in value:
                cell.fill = major_fill
            elif 'minor' in value:
                cell.fill = minor_fill
            elif 'low' in value:
                cell.fill = low_fill

    def _add_impact_formatting(self, sheet):
        """Add conditional formatting for impact column"""
        # Find impact column index
        impact_col = None
        for i, cell in enumerate(sheet[1]):
            if cell.value == 'Impact':
                impact_col = i + 1
                break
        
        if impact_col is None:
            return
        
        # Define styles
        high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")     # Red
        medium_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")   # Orange
        low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")      # Green
        
        # Apply styles
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=impact_col)
            value = cell.value.lower() if cell.value else ""
            
            if 'high' in value:
                cell.fill = high_fill
            elif 'medium' in value:
                cell.fill = medium_fill
            elif 'low' in value:
                cell.fill = low_fill

    def _add_category_pie_chart(self, sheet):
        """Add a pie chart for categories"""
        try:
            # Create pie chart
            chart = PieChart()
            chart.title = "Incidents by Category"
            
            # Get data range
            data_range = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
            categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            
            # Add data to chart
            chart.add_data(data_range)
            chart.set_categories(categories)
            
            # Set chart style
            chart.style = 10
            
            # Add the chart to the sheet
            sheet.add_chart(chart, "E2")
            
        except Exception as e:
            logger.error(f"Error adding pie chart: {str(e)}") 